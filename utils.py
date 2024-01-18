import os
import hashlib
import hmac
import base64
import urllib.parse as urlparse

import openpyxl
import xlsxwriter

from secret import secret

def sign_url(url):

    url = urlparse.urlparse(url)
    url_to_sign = url.path + "?" + url.query

    decoded_key = base64.urlsafe_b64decode(secret)
    signature = hmac.new(decoded_key, str.encode(url_to_sign), hashlib.sha1)
    encoded_signature = base64.urlsafe_b64encode(signature.digest())

    original_url = url.scheme + "://" + url.netloc + url.path + "?" + url.query
    return f"{original_url}&signature={encoded_signature.decode()}"

def load_exclude_ids(file):
    if not os.path.exists(file): return []

    list_of_ids = []
    with open(file, "rt") as file:
        lines = file.readlines()
        for line in lines:
            list_of_ids.append(line.split("\n")[0])
        file.close()

    return list_of_ids

def add_id_exclucion(ids, file):
    if not os.path.exists(file): open(file, "wt").close()

    with open(file, "at") as file:
        for id in ids:
            file.write(id+"\n")
        file.close()

def append_previous_excel(file):
    if not os.path.exists(file):
        workbook = xlsxwriter.Workbook('output.xlsx')
        worksheet = workbook.add_worksheet()
        rows = []
    else:
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]

        rows = []
        for row in sheet.values:
            rows.append(row)
        # Remove row with titles
        rows.pop(0)
        
        book.close()

        workbook = xlsxwriter.Workbook('output.xlsx')
        worksheet = workbook.add_worksheet()

    bold_format = workbook.add_format()
    bold_format.set_bold()

    worksheet.write("A1", "Osoite", bold_format)
    worksheet.set_column(0, 0, 25)
    worksheet.write("B1", "Kunta/Kaupunki", bold_format)
    worksheet.set_column(1, 1, 20)
    worksheet.write("C1", "Kartalla", bold_format)
    worksheet.write("D1", "Kuvaus", bold_format)
    worksheet.set_column(3, 3, 40)

    for i, row in enumerate(rows):
        try:
            worksheet.write(i+1, 0, row[0])
            worksheet.write(i+1, 1, row[1])
            worksheet.write_url(i+1, 2, row[2], string="linkki")
            worksheet.write(i+1, 3, row[3])
        except Exception as e:
            print("Jotain meni pieleen excelin kanssa :/")

    return workbook, worksheet, len(rows)+1